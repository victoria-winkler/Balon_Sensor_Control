"""
BAL.ON SENSOR TEST

Project for the Quality-Control Department

Device under Quality Control: BAL.ON E-Box and Sensors
For detailed information on the Software read the attached README.md

Victoria Winkler (v.winkler@trattereng.com)
Benjamin Stuppner (b.stuppner@trattereng.com)
"""
import tkinter
from functools import partial

"""

########         IMPORTS        ########

"""
from bleak import BleakClient, BleakScanner
import re
import numpy as np
from time import sleep
import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.animation as animation
from tkinter import *
from tkinter import ttk
import asyncio
import openpyxl
from openpyxl import Workbook
import os
import queue
from queue import Queue
import threading

"""

########         GLOBAL VARIABLES         ########

"""
#scan variables
the_scan_result_queue = Queue()
count_spinner = 0
internal_count = 0


#connect variables
weight_characteristic_uuid = "00004444-8e22-4541-9d4c-21edae82ed19"
data_write = "00008888-8e22-4541-9d4c-21edae82ed19"
value_to_write = "0401"
mac_address_connected = ''
name_connected = ''
serial_connected = 'myserial'       #TODO -> change dynamically
the_connect_queue = Queue()         #used to pass data from connect to update_graph


the_read_handler_queue = Queue()            # linked to the_graph_destroy_queue
the_read_handler_queue.put(2)


#graph variables
the_graph_handler_queue = Queue()
the_graph_handler_queue.put(0)

the_update_graph_queue = Queue()            #updates update_graph with values from BLE
count_graph = 0

x_array = []
y1_array = []
y2_array = []
y3_array = []
y4_array = []
y5_array = []
y6_array = []
y7_array = []
y8_array = []
y9_array = []
timestamp = []

"""

########         SCANNER         ########

"""


def command_start_scan():
    threading.Thread(target=wrap_async_start_scan).start()
    root.after(100, handle_scan)


def wrap_async_start_scan():
    asyncio.run(scanner_thread())


async def scanner_thread():
    devices = await BleakScanner.discover(timeout=5)

    # has devices
    if len(devices) > 0:
        print('has devices')     #todo
        # searches for a BAL.ON device
        for device in devices:
            string_device_name = device.name
            print(device)    #todo

            search_balon = re.search("^BAL.ON", str(string_device_name))

            if search_balon:
                the_scan_result_queue.put(device)
                break

    else:
        print('no devices')      #todo

    the_scan_result_queue.put(None)
    return


def handle_scan():

    # SPINNING DOTS
    global count_spinner
    global internal_count

    if count_spinner == 1:
        status_bar_label.config(text='. ' * internal_count, font=('Helvetica', 24))

    count_spinner += 1
    if count_spinner == 4:
        internal_count += 1
        count_spinner = 0

        if internal_count == 4:
            internal_count = 0


    # actual callback PART
    try:
        device = the_scan_result_queue.get(block=False)
    except queue.Empty:
        root.after(100, handle_scan)
        return

    if device is not None:
        print('after_callback got', device.address)
        address_entry.insert(0, str(device.address))
        check_mac_addresses(device.address)
        connect_button.state(["!disabled"])

    if device is None:
        status_bar_label.config(text='NO DEVICE FOUND, ensure the device is Bluetooth enabled and scan again', font=('Helvetica', 12))






"""

########         VALIDATE SCANNED DEVICE         ########

"""

def check_mac_addresses(address):
    # VERIFY IF FILE EXISTS
    path = os.getcwd()
    excel_file = f'{path}/mac_addresses.xlsx'
    file_exists = os.path.isfile(excel_file)

    # LOAD FILE
    if file_exists:
        print("file exist")

        file_name = 'mac_addresses.xlsx'

        wb = openpyxl.load_workbook(file_name)
        # load worksheet
        sheet = wb['MAC ADDRESSES']  # load worksheet
        # store all addresses in list
        first_column = sheet['A']
        used_mac_addresses = [cell.value for cell in first_column[0:]]

        print(used_mac_addresses)

        # check if address is present in the list
        if address in used_mac_addresses:
            status_bar_label.config(text=f'Device with this {address} has already been used. Please Scan again or verify address using an NFC scanner', background='#880808', font=('Helvetica', 16))
            # label_address['text'] = f'Device with this {address} has already been used'
            print(f'{address} is present in the list')
        else:
            print(f'{address} is not present in the list')
            status_bar_label.config(text=f'Device with this {address} has not been used yet', background='#008000', font=('Helvetica', 16))
            # append address
            current_time = datetime.datetime.now()
            data = (f'{address}', str(current_time))
            sheet.append(data)


    # CREATE FILE + ADD FIRST ADDRESS
    else:
        status_bar_label.config(text=f'Device with this {address} has not been used yet', background='#008000', font=('Helvetica', 16))
        print("file doesn't exist")
        wb = Workbook()
        sheet = wb.create_sheet("MAC ADDRESSES", 0)

        current_time = datetime.datetime.now()

        # enter first address
        sheet['A1'] = f'{address}'
        sheet['B1'] = f'{current_time}'

    wb.save("mac_addresses.xlsx")




"""

########         CONNECTION + SUBSCRIPTION        ########

"""

def command_connect():
    threading.Thread(target=wrap_async_connect).start()
    scan_button.state(["disabled"])



def wrap_async_connect():
    asyncio.run(thread_connect())



# TODO
async def thread_connect():
    mac_address = address_entry.get()

    # CONNECTION
    try:
        connected_device = await BleakScanner.find_device_by_address(mac_address, timeout=20)
        status_bar_label['text'] = f'connected to {mac_address}'
        async with BleakClient(connected_device) as client:

            ### GET SERIAL NUMBER OF DEVICE - todo ask ROLAND FOR subscription string


            # SUBSCRIPTION
            write_value = bytearray([4, 1])
            await client.write_gatt_char(data_write, write_value)
            print("subscribed successfully")
            status_bar_label['text'] = "subscribed"

            global mac_address_connected
            mac_address_connected = mac_address

            #disable
            connect_button.state(["disabled"])

            #enable
            start_reading_button.state(["!disabled"])

            alive_loop = 1

            while alive_loop == 1:
                try:
                    handler = the_read_handler_queue.get(block=False)
                    print(f'this is READ Handler: {handler}')

                    # READ DATA
                    if handler == 0:
                        data = await client.read_gatt_char(weight_characteristic_uuid)          #reads data
                        decoded_data = encode_bytes_to_string(data)
                        # print(decoded_data)
                        the_update_graph_queue.put(decoded_data)
                        # the_connect_queue.put(decoded_data) #might not be necessary -> pass to graph update

                        sleep(0.3)  # TODO might be necessary for better results!!
                        the_read_handler_queue.put(0)


                    # do not read data
                    if handler == 2:
                        print('data on hold_2')

                        sleep(0.3)  # TODO might be necessary for better results!!
                        # the_read_handler_queue.put(2)

                    # disconnect
                    if handler == 1:
                        alive_loop = 0

                except queue.Empty:
                    print('Error empty read handler queue')
                    sleep(0.3)

    except Exception as e:
        status_bar_label['text'] = f'unable to connect to {mac_address}'
        print(f"Could not connect to device")
        print(f"Error: {e}")

    status_bar_label['text'] = "disconnected"
    mac_address_connected = ''
    return



def encode_bytes_to_string(bytes_):
    # Create a template for how to slice up the bytes
    data_segments = np.dtype([('battery_level', np.uint8), ('is_charging', np.uint8)])
    formatted_data = np.frombuffer(bytes_, dtype=data_segments)
    return formatted_data




"""

########         GRAPH         ########

"""


def graph_update(i, line, line2, line3, line4, line5, line6, line7, line8, line9):
    # global count_graph
    # print(f' - {i} vs. graph update count: {count_graph}')

    try:
        data = the_update_graph_queue.get(block=False)
    except queue.Empty:
        print("NO DATA")
        return

    input_data_list = data.tolist()

    sensor_1 = input_data_list[3][0]
    sensor_2 = input_data_list[4][0]
    sensor_3 = input_data_list[5][0]
    sensor_4 = input_data_list[6][0]
    sensor_5 = input_data_list[7][0]
    sensor_6 = input_data_list[8][0]
    sensor_7 = input_data_list[9][0]
    sensor_8 = input_data_list[3][1]
    sensor_9 = input_data_list[4][1]

    # ACTUAL CODE
    y1_array.append(sensor_1)
    y2_array.append(sensor_2)
    y3_array.append(sensor_3)
    y4_array.append(sensor_4)
    y5_array.append(sensor_5)
    y6_array.append(sensor_6)
    y7_array.append(sensor_7)
    y8_array.append(sensor_8)
    y9_array.append(sensor_9)

    x_array.append(i)
    line.set_data(x_array, y1_array)
    line2.set_data(x_array, y2_array)
    line3.set_data(x_array, y3_array)
    line4.set_data(x_array, y4_array)
    line5.set_data(x_array, y5_array)
    line6.set_data(x_array, y6_array)
    line7.set_data(x_array, y7_array)
    line8.set_data(x_array, y8_array)
    line9.set_data(x_array, y9_array)

    current_time = datetime.datetime.now()
    timestamp.append(current_time)

    # count_graph = count_graph + 1


def graph_thread():
    threading.Thread(target=create_graph).start()

#TODO
def create_graph():
    print('create graph')

    graph_frame.pack_forget()   #needs to be here -> otherwise it won't display the graph

    line, = ax.plot(x_array, y1_array, label="sensor_1")
    line2, = ax.plot(x_array, y2_array, label="sensor_2")
    line3, = ax.plot(x_array, y3_array, label="sensor_3")
    line4, = ax.plot(x_array, y4_array, label="sensor_4")
    line5, = ax.plot(x_array, y5_array, label="sensor_5")
    line6, = ax.plot(x_array, y6_array, label="sensor_6")
    line7, = ax.plot(x_array, y7_array, label="sensor_7")
    line8, = ax.plot(x_array, y8_array, label="sensor_8")
    line9, = ax.plot(x_array, y9_array, label="sensor_9")


    ax.legend()

    canvas.draw()
    graph_frame.pack()

    ani_test = animation.FuncAnimation(fig, partial(graph_update, line=line, line2=line2, line3=line3, line4=line4, line5=line5, line6=line6, line7=line7, line8=line8, line9=line9), cache_frame_data=False, interval=300, blit=False)

    alive_loop = 1

    while alive_loop == 1:
        try:
            destroy = the_graph_handler_queue.get(block=False)

            if destroy == 0:            #graph is running fine
                the_graph_handler_queue.put(0)

            if destroy == 2:            #puts graph on hold
                ani_test.pause()

            if destroy == 1:               #destroys current graph -> resets
                _clear_canvas()
                ax.clear()
                plt.close(fig)
                graph_frame.pack_forget() #todo check if this is better!
                ax.set_ylim(0, 300)
                ax.set_xlim(0, 100)

                alive_loop = 0
                the_graph_handler_queue.put(0)

                global count_graph
                count_graph = 0

                x_array.clear()
                y1_array.clear()
                y2_array.clear()
                y3_array.clear()
                y4_array.clear()
                y5_array.clear()
                y6_array.clear()
                y7_array.clear()
                y8_array.clear()
                y9_array.clear()

        except queue.Empty:
            print('nothing in graph destroyer queue')

        sleep(0.3)

    return





"""

########         STORE DATA         ########

"""

# new simpler store data function -> does not create subfolder
def store_data_windows(address, serial):
    folder_path = os.getcwd()

    print(f'path exists: {folder_path}')

    file_name = f'{address}_{serial}.xlsx'
    # check if file with data exists: mac_serial_xx
    excel_file = f'{folder_path}/{file_name}'
    file_exists = os.path.isfile(excel_file)


    # create file
    if not file_exists:
        print('file does not exist')
        wb = Workbook()
        # wb.create_sheet("version_1", 0)
        wb.save(f'{folder_path}/{file_name}')


    # LOAD FILE
    wb = openpyxl.load_workbook(f'{folder_path}/{file_name}')
    all_sheets = wb.sheetnames
    print(all_sheets)
    number = len(all_sheets)
    print(number)

    sheet = wb.create_sheet(f'version_{number}', 0)
    print(sheet)

    # store data: sensors1 ...sensor9 + timestamp
    amount = len(y1_array)
    i = 0

    for i in range(amount):
        data = (y1_array[i], y2_array[i], y3_array[i], y4_array[i], y5_array[i], y6_array[i], y7_array[i], y8_array[i], y9_array[i], timestamp[i])
        sheet.append(data)

    # close
    wb.save(f'{folder_path}/{file_name}')




def store_data(address, serial):
    current_path = os.getcwd()
    folder_path = f'{current_path}/sensor_data'

    # check if folder exists -> else create
    path_exists = os.path.exists(folder_path)

    # path doesn't exist -> create
    if not path_exists:
        print('path doesnt exist')
        directory = 'sensor_data'
        path = os.path.join(current_path, directory)
        os.mkdir(path)

    print(f'path exists: {folder_path}')

    file_name = f'{address}_{serial}.xlsx'
    # check if file with data exists: mac_serial_xx
    excel_file = f'{folder_path}/{file_name}'
    file_exists = os.path.isfile(excel_file)


    # create file
    if not file_exists:
        print('file does not exist')
        wb = Workbook()
        # wb.create_sheet("version_1", 0)
        wb.save(f'{folder_path}/{file_name}')


    # LOAD FILE
    wb = openpyxl.load_workbook(f'{folder_path}/{file_name}')
    all_sheets = wb.sheetnames
    print(all_sheets)
    number = len(all_sheets)
    print(number)

    sheet = wb.create_sheet(f'version_{number}', 0)
    print(sheet)

    # store data: sensors1 ...sensor9 + timestamp
    amount = len(y1_array)
    i = 0

    for i in range(amount):
        data = (y1_array[i], y2_array[i], y3_array[i], y4_array[i], y5_array[i], y6_array[i], y7_array[i], y8_array[i], y9_array[i], timestamp[i])
        sheet.append(data)

    # close
    wb.save(f'{folder_path}/{file_name}')



"""

########         BUTTON FUNCTIONS         ########

"""

# command -> 0
def command_start_reading():
    print('start GRAPH - start reading from BLE')
    print(f'XARRAY: {len(x_array)}')

    # disable
    start_reading_button.state(["disabled"])


    # enable
    stop_reading_button.state(["!disabled"])
    # disconnect_button.state(["!disabled"])

    graph_thread()
    the_read_handler_queue.put(0)
    the_graph_handler_queue.put(0)




# command -> 2
def command_stop_reading():
    print('stop graph command -  stops reading from BLE device')

    # disable
    stop_reading_button.state(["disabled"])


    # enable
    # disconnect_button.state(["!disabled"])
    # delete_current_button.state(["!disabled"])
    store_data_button.state(["!disabled"])

    the_read_handler_queue.put(2)
    the_graph_handler_queue.put(2)


#command -> 1 USED TO BE "DESTROY"
#command -> 2 for READ
def command_delete_current():
    print('prepares for rereading')

    # disable
    # delete_current_button.state(["disabled"])
    store_data_button.state(["disabled"])

    # enable
    start_reading_button.state(["!disabled"])

    the_graph_handler_queue.put(1)
    the_read_handler_queue.put(2)


# command -> 1 for BOTH
def command_disconnect():
    print('disconnecting from BLE device')

    # disable
    # disconnect_button.state(["disabled"])


    # enable
    scan_button.state(["!disabled"])
    connect_button.state(["!disabled"])

    the_read_handler_queue.put(1)
    the_graph_handler_queue.put(1)



def command_store_data():
    print('store data')

    # disable
    store_data_button.state(["disabled"])

    store_data_windows(mac_address_connected, serial_connected)

    # enable
    # delete_current_button.state(["!disabled"])
    # disconnect_button.state(["!disabled"])




"""

########         GUI         ########

"""

root = Tk()
root.geometry('1200x500')
root.title('BAL.ON SENSOR TEST')
root.resizable(False, False)

# style = ttk.Style()
# style.map("Mod.TButton", background=[("active", "red"), ("!active", "blue")])


container_main = ttk.Frame(root)
container_main.pack(side='left', anchor='n')

# LEFT SIDE

left_container = ttk.Frame(container_main)
left_container.pack(side='left', padx=20, pady=20)

inside1_left = ttk.Frame(left_container)
inside1_left.pack(side='top', pady=10)

scan_button = ttk.Button(inside1_left, text="SCAN", style="Mod.TButton", command=command_start_scan)
scan_button.pack(side='left', padx=20)

address_entry = ttk.Entry(inside1_left)
address_entry.pack(side='right', padx=20)

connect_button = ttk.Button(left_container, text="CONNECT", command=command_connect)
connect_button.pack(pady=10)
connect_button.state(["disabled"])

start_reading_button = ttk.Button(left_container, text="START READING", command=command_start_reading)
start_reading_button.pack(pady=10)
start_reading_button.state(["disabled"])

stop_reading_button = ttk.Button(left_container, text="STOP READING", command=command_stop_reading)
stop_reading_button.pack(pady=10)
stop_reading_button.state(["disabled"])

# ReRead + Store

inside2_left = ttk.Frame(left_container)
inside2_left.pack(side='top', pady=10)

# delete_current_button = ttk.Button(inside2_left, text="RE-READ", command=command_delete_current)
# delete_current_button.pack(side='left', padx=20)
# delete_current_button.state(["disabled"])

store_data_button = ttk.Button(inside2_left, text="STORE DATA", command=command_store_data)
store_data_button.pack(side='right', padx=20)
store_data_button.state(["disabled"])

# DISCONNECT + Status Bar

# disconnect_button = ttk.Button(left_container, text="DISCONNECT", command=command_disconnect)
# disconnect_button.pack(pady=10)
# disconnect_button.state(["disabled"])

status_bar_label = ttk.Label(left_container, font=('Helvetica', 24), wraplength=300)
status_bar_label.pack(side='bottom', pady=10, anchor='s')

# RIGHT SIDE

right_container = ttk.Frame(container_main)
right_container.pack(padx=20, pady=20)

device_information = ttk.Label(right_container, text="this is a label")
device_information.pack(anchor='n', side='top', pady=10)

inside1_right = ttk.Frame(container_main)
inside1_right.pack(side='top', pady=10)

placeholder_frame = ttk.Frame(inside1_right, width=10)
placeholder_frame.pack(side='left', padx=20, expand=True)



#GRAPH
graph_frame = tkinter.Frame(inside1_right)

fig = plt.figure(figsize=(14, 4.5), dpi=100)

ax = fig.add_subplot(1, 1, 1)
ax.set_ylim(0, 300)
ax.set_xlim(0, 100)

canvas = FigureCanvasTkAgg(fig, master=graph_frame)
canvas.get_tk_widget().pack(side='right', fill='both', expand=1)
graph_frame.pack()      #todo might cause an issue!!


# HELPER METHODS FOR GUI

def _clear_canvas():
    for item in canvas.get_tk_widget().find_all():
        canvas.get_tk_widget().delete(item)


root.mainloop()
